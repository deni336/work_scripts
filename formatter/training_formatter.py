import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from fuzzywuzzy import process

def standardize_course_titles(course_titles):
    standardized_titles = {}
    
    for title in course_titles:
        if not standardized_titles:
            standardized_titles[title] = title
            continue
        
        best_match = process.extractOne(title, standardized_titles.keys())
        
        if best_match and best_match[1] > 85:
            standardized_titles[title] = standardized_titles[best_match[0]]
        else:
            standardized_titles[title] = title
    
    return standardized_titles

def format_envision_export(input_file, output_file):
    df = pd.read_excel(input_file)
    
    column_mappings = {
        "Airman Name": "Name",
        "Office Symbol": "Office",
        "Task Title": "Course Title",
        "Due Date": "Due Date",
        "Task Name": "Task Name",
        "Task Status": "Status",
        "Task Type": "Type",
        "Is Next 30": "Next 30 Days"
    }
    
    df.rename(columns=column_mappings, inplace=True)
    df = df[df["Type"] == "CBT"]
    
    df["Course Title"] = df["Course Title"].str.replace(r"99ARS", "", regex=True)
    df["Course Title"] = df["Course Title"].str.replace(r"for .*", "", regex=True)
    df["Course Title"] = df["Course Title"].apply(lambda x: x if '-' in x and (x.startswith('-') or x.endswith('-')) else x.replace('-', ''))
    
    exclude_courses = ["Comprehensive Airman Fitness"]
    df = df[~df["Course Title"].str.strip().isin(exclude_courses)]
    
    unique_titles = df["Course Title"].dropna().unique()
    title_mapping = standardize_course_titles(unique_titles)
    df["Course Title"] = df["Course Title"].map(title_mapping)
    
    df["Status"] = df["Status"].replace("MISSING", "Not Attempted")
    df.loc[df["Due Date"].isna() & (df["Status"] == "Not Attempted"), "Due Date"] = "Not Attempted"
    
    # Ensure Due Date is properly formatted
    df.loc[df["Due Date"] != "Not Attempted", "Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce").dt.strftime("%m/%d/%Y")
    
    df_pivot = df.pivot_table(index=["Name", "Office"], 
                              columns="Course Title", 
                              values="Due Date", 
                              aggfunc="first")
    
    min_participants = 5
    valid_columns = df_pivot.count() >= min_participants
    df_pivot = df_pivot.loc[:, valid_columns]
    
    df_pivot.reset_index(inplace=True)
    df_pivot.to_excel(output_file, index=False, engine='openpyxl')
    
    wb = load_workbook(output_file)
    ws = wb.active
    
    percentage_row = ["Percentage Completed", ""]
    for col_idx in range(3, ws.max_column + 1):  # Start at column C
        col_letter = get_column_letter(col_idx)
        
        formula = f'=IF(COUNTA({col_letter}3:{col_letter}{ws.max_row+1})=0, "0%", ' \
                  f'TEXT(COUNTIFS({col_letter}3:{col_letter}{ws.max_row+1}, ">"&TODAY()) / ' \
                  f'COUNTIFS({col_letter}3:{col_letter}{ws.max_row+1}, ">*") * 100, "0.00%"))'
        percentage_row.append(formula)
    
    ws.insert_rows(1)
    for col_idx, value in enumerate(percentage_row, start=1):
        ws.cell(row=1, column=col_idx, value=value)
    
    header_font = Font(bold=True)
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(horizontal="center")
        
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    today = datetime.today().date()
    thirty_days_out = today + timedelta(days=30)
    
    for row in ws.iter_rows(min_row=3, min_col=3):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = "MM/DD/YYYY"  # Ensure date formatting
                if cell.value.date() < today:
                    cell.fill = red_fill
                elif today <= cell.value.date() <= thirty_days_out:
                    cell.fill = yellow_fill
    
    wb.save(output_file)
    print(f"Formatted file saved as {output_file}")

# Example Usage
input_file = "Envision Training Export.xlsx"
output_file = "Formatted_Envision_Training.xlsx"
format_envision_export(input_file, output_file)
